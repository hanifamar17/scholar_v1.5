from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    jsonify,
    session,
    flash
)
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mysqldb import MySQL
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime as dt
import mysql.connector
import MySQLdb
import subprocess
import datetime
import json
import pandas as pd
import io
import os
import math
import pdfkit
from dotenv import load_dotenv


load_dotenv()
app = Flask(__name__, static_url_path="/static")
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
CORS(app)

app.config["MYSQL_HOST"] = "127.0.0.1"
app.config["MYSQL_USER"] = "root"
app.config["MYSQL_PASSWORD"] = ""
app.config["MYSQL_DB"] = "publications"
#BACKUP_PATH = 'C:\\Users\\WINDOWS\\Downloads\\backups'

mysql = MySQL(app)

#GLOBAL PATH
MYSQLDUMP_PATH = r'E:\App-Development\1-Tools-Library-Environment\laragon\bin\mysql\mysql-8.0.30-winx64\bin\mysqldump.exe'
WKHTMLTOPDF_PATH= r'E:\\App-Development\\1-Tools-Library-Environment\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
FLASK_PORT=5000
    
#ERROR HANDLING GLOBAL
@app.errorhandler(MySQLdb.OperationalError)
def handle_db_connection_error(e):
    error_msg = "Oops! It looks like there's a problem with the server. Please check if the database server is running properly."
    return render_template("error.html", error=error_msg), 500

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    error_msg = "Terjadi kesalahan internal pada server."
    return render_template("error.html", error=error_msg), 500

#BACKUP-RESTORE
# Konfigurasi backup directory
BACKUP_DIR = os.path.join(app.root_path, 'static', 'backups')
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

BACKUP_JSON = os.path.join('static', 'backups_json', 'backups.json')
def load_backup_data():
    # Memuat data dari backup.json jika file ada
    if os.path.exists(BACKUP_JSON):
        with open(BACKUP_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_backup_data(backups):
    # Menyimpan data backup ke backup.json
    with open(BACKUP_JSON, 'w') as f:
        json.dump(backups, f, indent=4)

def file_size_convert(size_in_bytes):
    if size_in_bytes < 1024:
        return f"{size_in_bytes} Bytes"
    elif size_in_bytes < 1048576:
        return f"{size_in_bytes / 1024:.2f} KB"
    elif size_in_bytes < 1073741824:
        return f"{size_in_bytes / 1048576:.2f} MB"
    else:
        return f"{size_in_bytes / 1073741824:.2f} GB"

# Endpoint untuk menampilkan halaman backup dan restore
@app.route('/backup-restore')
def backup_restore():
    # Ambil daftar file backup
    backups = load_backup_data()
    for filename in os.listdir(BACKUP_DIR):
        if filename.endswith('.sql'):
            filepath = os.path.join(BACKUP_DIR, filename)
            timestamp = os.path.getmtime(filepath)
            date = dt.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            size = os.path.getsize(filepath)
            readable_size = file_size_convert(size)
            
            if not any(b['filename'] == filename for b in backups):
                backups.append({
                    'filename': filename,
                    'date': date,
                    'size': readable_size
                })          
    
    return render_template('/backup/backup-restore.html', backups=backups)

# Endpoint untuk melakukan backup database
@app.route('/backup', methods=['POST', 'GET'])
def backup():
    timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_{timestamp}.sql'
    filepath = os.path.join(BACKUP_DIR, filename)

    # Perintah mysqldump (gunakan path lengkap)
    mysqldump_cmd = [
        MYSQLDUMP_PATH,
        '-h', app.config["MYSQL_HOST"],
        '-u', app.config["MYSQL_USER"],
        app.config['MYSQL_DB']
    ]

    # Lakukan backup menggunakan subprocess
    try:
        with open(filepath, 'w') as f:
            subprocess.run(mysqldump_cmd, stdout=f, check=True)
        
        if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
            flash('Backup berhasil dibuat', 'success')

            # Menambahkan informasi backup baru ke dalam backup.json
            backups = load_backup_data()  # Load existing backups
            readable_size = file_size_convert(os.path.getsize(filepath))
            backups.append({
                'filename': filename,
                'date': dt.now().strftime('%Y-%m-%d %H:%M:%S'),
                'size': readable_size
            })
            save_backup_data(backups)  # Simpan data yang diperbarui ke dalam backup.json
        else:
            raise Exception("Backup file is empty")
    except subprocess.CalledProcessError as e:
        flash(f"Backup error: {str(e)}", 'error')
        print(f"Backup error: {str(e)}")
    except Exception as e:
        flash(f"Backup error: {str(e)}", 'error')
        print(f"Backup error: {str(e)}")

    return redirect(url_for('backup_restore'))

# Endpoint untuk mengunduh file backup
@app.route('/backup/download/<filename>')
def download(filename):
    return send_file(os.path.join(BACKUP_DIR, filename), as_attachment=True)

# Endpoint untuk merestore database dari file .sql
@app.route('/restore', methods=['POST'])
def restore():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('backup_restore'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('backup_restore'))
    
    if file and file.filename.endswith('.sql'):
        filepath = os.path.join(BACKUP_DIR, file.filename)
        file.save(filepath)

        # Perintah mysql untuk restore (gunakan path lengkap)
        mysql_cmd = [
            r'E:\App-Development\1-Tools-Library-Environment\laragon\bin\mysql\mysql-8.0.30-winx64\bin\mysql.exe',  # Path ke mysql
            '-h', app.config["MYSQL_HOST"],
            '-u', app.config["MYSQL_USER"],
            f'-p{app.config["MYSQL_PASSWORD"]}',
            app.config['MYSQL_DB']
        ]

        # Lakukan restore menggunakan subprocess
        try:
            with open(filepath, 'r') as f:
                subprocess.run(mysql_cmd, stdin=f, check=True)
            flash('Database berhasil di-restore', 'success')
        except subprocess.CalledProcessError as e:
            flash(f"Restore error: {str(e)}", 'error')
            print(f"Restore error: {str(e)}")
        except Exception as e:
            flash(f"Restore error: {str(e)}", 'error')
            print(f"Restore error: {str(e)}")

    return redirect(url_for('backup_restore'))

#LOGIN
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id_admin, username, password):
        self.id_admin = id_admin
        self.username = username
        self.password = password
    
    def get_id(self):
        return str(self.id_admin)

@login_manager.user_loader
def load_user(id_admin):
    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM admin WHERE id_admin = %s', (id_admin,))
        account = cursor.fetchone()
        if account:
            return User(account['id_admin'], account['username'], account['password'])
        return None
    except MySQLdb.OperationalError as e:
        app.logger.error(f"Database connection failed: {e}")
        return None

@app.route('/', methods=['GET', 'POST'])
def login():
    try:
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']

            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT * FROM admin WHERE username = %s', (username,))
            account = cursor.fetchone()

            if account and account['password'] == password:  # Pastikan untuk menggunakan hash password di produksi
                user = User(account['id_admin'], account['username'], account['password'])
                login_user(user)
                return redirect(url_for('index'))
            else:
                flash('Invalid username or password', 'danger')
        return render_template('login.html')
    except MySQLdb.OperationalError as e:
        error_msg = "Terjadi masalah saat mengakses sistem. Pastikan Database server anda sudah dijalankan."
        return render_template("error-login.html", error=error_msg)

#@app.route("/testing", methods=["GET", "POST"])
#def testing():
#    if request.method == "POST":
#        selected_articles = request.form.getlist("selected_articles")
#        selected = ", ".join(selected_articles)
#
#        # Format string pengembalian untuk debug
#        return f"Received author_id: {selected}"

@app.route("/dashboard")
@login_required
def index():
    return render_template("index.html")
    
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

#ADMIN
@app.route("/admin")
def admin():
    cur = mysql.connection.cursor()
    cur.execute("""SELECT * FROM admin""")
    data_admin = cur.fetchall()
    cur.close()
    return render_template(
        "admin/admin.html",
        admin=data_admin
    )

@app.route("/admin/add", methods=["GET", "POST"])
def admin_add():
    return render_template("admin/add.html")

@app.route("/admin/add/submit", methods=["GET", "POST"])
def admin_add_submit():
    username = request.form["username"]
    password = request.form["password"]

    cur = mysql.connection.cursor()
    cur.execute(
        "INSERT INTO admin (username, password) VALUES (%s, %s)", (username, password)
    )
    mysql.connection.commit()
    cur.close()

    flash('Admin berhasil ditambahkan.', 'success')
    return redirect(url_for("admin"))

@app.route("/admin/delete/<int:id_admin>", methods=["POST", "DELETE"])
def admin_delete(id_admin):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM admin WHERE id_admin = %s", (id_admin,))
    mysql.connection.commit()
    cur.close()

    flash('Admin berhasil dihapus.', 'success')
    return redirect(url_for("admin"))

@app.route("/admin/update/form/<int:id_admin>", methods=["GET", "POST"])
def admin_update_form(id_admin):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM admin WHERE id_admin = %s", (id_admin,))
    data_admin = cur.fetchone()
    cur.close()

    return render_template("admin/update-form.html", admin=data_admin)

@app.route("/admin/update/<int:id_admin>", methods=["POST", "GET"])
def admin_update(id_admin):
    username = request.form["username"]
    password = request.form["password"]

    cur = mysql.connection.cursor()
    cur.execute(
        "UPDATE admin SET username = %s, password = %s WHERE id_admin = %s",
        (username, password, id_admin),
    )
    mysql.connection.commit()
    cur.close()

    flash('Admin berhasil diperbarui.', 'success')
    return redirect(url_for("admin"))

#PENCARIAN
@app.route("/authors")
def authors():
    with open("output_profiles.json", "r") as profiles:
        data_profiles_json = json.load(profiles)
    return render_template("pencarian/pencarian.html", profiles=data_profiles_json)
    
@app.route("/authors/get_authors", methods=["GET", "POST"])
def get_authors():
    query_profiles = request.form.get("query_profiles").lower()
    spider_name = "dataCrawler_4"

    subprocess.run(["scrapy", "crawl", spider_name, "-o", "output_profiles.json", "-a", f"query_keyword={query_profiles}",])

    with open("output_profiles.json") as profiles:
        data_profiles = json.load(profiles)

    session['data_profiles'] = data_profiles
    session['query_profiles'] = query_profiles

    return redirect(url_for("get_profiles", query=query_profiles))

@app.route("/authors/get_articles", methods=["GET", "POST"])
def get_articles():
    if request.method == "POST":
        selected_query = request.form.getlist("selected_authors")
        selected_query = ",".join(selected_query)
        # return f"Received author name: {selected_query}"
        spider_name = "dataCrawler_3"
        subprocess.run(["scrapy", "crawl", spider_name, "-o", "output.json", "-a", f"author_id={selected_query}",])

        with open("output.json", encoding='utf-8') as items_file:
            data_articles = json.load(items_file)

        with open("output_profiles.json") as profiles:
            data_profiles_json = json.load(profiles)
        
        #return render_template(
        #    "pencarian/get_profiles.html", output=data_articles, profiles=data_profiles_json
        #)
        return redirect(url_for('results'))
    else:
        selected_query = request.args.get("selected_author")
        # return f"Received author name: {selected_query}"
        spider_name = "dataCrawler_3"
        subprocess.run(["scrapy", "crawl", spider_name, "-o", "output.json", "-a", f"author_id={selected_query}",])

        with open("output.json", encoding='utf-8') as items_file:
            data_articles = json.load(items_file)

        with open("output_profiles.json") as profiles:
            data_profiles_json = json.load(profiles)
        
        #return render_template(
         #   "pencarian/get_profiles.html", output=data_articles, profiles=data_profiles_json
        #)
        return redirect(url_for('results'))

@app.route("/authors/get_profiles")
def get_profiles():
    query_param = request.args.get('query', '-')
    query_profiles = [query.strip() for query in query_param.split(',')]

    #query_profiles = session.get("query_profiles", "-")
    data_profiles = session.get("data_profiles", "-")
    not_found_queries = []

    try:
        with open("output_profiles.json") as profiles:
            data_profiles_json = profiles.read()
            data_profiles_json = json.loads(data_profiles_json)

            names = [profile.get('name', '').lower() for profile in data_profiles_json]
            for query in query_profiles:
                if not any(query.lower() in name for name in names):
                    not_found_queries.append(query)

            if not_found_queries:                
                raise ValueError(f"Tidak ditemukan hasil pencarian untuk")

            return render_template("pencarian/get_profiles.html", profiles=data_profiles)
    
    except ValueError as ve:
        not_found_queries_str = ', '.join(not_found_queries)
        return render_template('pencarian/get_profiles.html', error_message=str(ve), query_profiles=not_found_queries_str, profiles=data_profiles)
    
@app.route("/authors/history_get_articles")
def history_get_articles():
    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT query, GROUP_CONCAT(thumbnail) AS thumbnail, MAX(updated_at) AS updated_at "
        "FROM publikasi "
        "GROUP BY query "
        "ORDER BY updated_at DESC"
    )
    authors = cur.fetchall()
    author_data = []
    for author in authors:
        query = author[0].lower()
        thumbnail = author[1]
        updated_at = author[2]
        updated_at = updated_at.strftime("%d-%m-%Y %H:%M:%S")
        cur.execute(
            "SELECT prodi FROM dosen WHERE LOWER(nama_dosen) = %s", (query,)
        )
        result = cur.fetchone()
        if result:
            prodi = result[0]
        else:
            prodi = "-"
        author_data.append(
            {
                "query": query,
                "thumbnail": thumbnail,
                "prodi": prodi,
                "updated_at": updated_at,
            }
        )
    cur.close()
    return render_template("pencarian/history_get_articles.html", history=author_data)
    

@app.route("/authors/history_profiles")
def history_profiles():
    cur = mysql.connection.cursor()
    cur.execute("SELECT name, affiliations, thumbnail, author_id, interests FROM profiles ORDER BY updated_at DESC")
    history_profiles = cur.fetchall()
    converted_profiles = []
    for profile in history_profiles:
        name, affiliations, thumbnail, author_id, interests = profile
        # Jika 'interests' adalah string JSON, konversi ke list Python
        if isinstance(interests, str):
            interests = json.loads(interests)  # Mengubah string JSON jadi list
        converted_profiles.append((name, affiliations, thumbnail, author_id, interests))
    cur.close()
    return render_template("pencarian/history_profiles.html", history_profiles=converted_profiles)

#HASIL PENCARIAN
#@app.route("/results", methods=["GET", "POST"])
#def results():
#    page = request.args.get("page", 1, type=int)
#
#    if request.method == "POST":
#        per_page = int(request.form.get("per_page", "10"))
#        return redirect(url_for('results', per_page=per_page, page=1))
#    else:
#        per_page = int(request.args.get("per_page", "10"))
#
#    with open("output.json", 'r', encoding='utf-8') as f:
#        data = json.load(f)
#
#     # Ganti nilai None dengan 0
#    for item in data:
#        for key, value in item.items():
#            if value is None:
#                item[key] = 0
#
#    # Dapatkan semua query yang unik
#    queries = set(item["query"].lower() for item in data if "query" in item)
#    queries = list(queries)
#
#    # for entry in data:
#    #   entry['title'] = entry['title'][0] if entry['title'] else ""
#
#    data_sorted = sorted(data, key=lambda x: x["query"])    
#
#    # Tentukan jumlah data per halaman
#    total = len(data_sorted)
#    total_pages = math.ceil(total / per_page)
#
#    # Tentukan batasan data untuk halaman saat ini
#    start = (page - 1) * per_page
#    end = start + per_page
#    paginated_data = data_sorted[start:end]
#
#    return render_template(
#        "results/results.html",
#        value=paginated_data,
#        queries=queries,
#        page=page,
#        per_page=per_page,
#        total_pages=total_pages,
#    )

@app.route("/results", methods=["GET", "POST"])
def results():
    if request.args.get('draw'):
        with open("output.json", 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Ganti nilai None dengan 0
        for item in data:
            for key, value in item.items():
                if value is None:
                    item[key] = 0
        
        draw = request.args.get('draw')
        start = int(request.args.get('start', 0))  # start point
        length = int(request.args.get('length', 10))  # number of records per page
        search_value = request.args.get('search[value]', '')  # search term

        # Filter data berdasarkan search_value
        if search_value:
            filtered_data = [
                item for item in data 
                if search_value in str(item.get("query", "")).lower() or
                   search_value in str(item.get("author", "")).lower() or
                   search_value in str(item.get("title", "")).lower() or
                   search_value in str(item.get("title_by_url", "")).lower() or
                   search_value in str(item.get("cited_by_value", "")).lower() or
                   search_value in str(item.get("cited_by_url", "")).lower() or
                   search_value in str(item.get("publication_year", "")).lower() or
                   search_value in str(item.get("publication", "")).lower()
            ]
        else:
            filtered_data = data
        
        filtered_data_sort = sorted(filtered_data, key=lambda x: x["query"])

        # Pagination (limit and offset)
        page_data = filtered_data_sort[start:start + length]

        results = [{
                "no": start + i + 1,
                "query": row.get("query", ""),
                "author": row.get("author", ""),
                "title": row.get("title", ""),
                "title_url": row.get("title_url", ""),
                "cited_by_value": row.get("cited_by_value", ""),
                "cited_by_url": row.get("cited_by_url", ""),
                "publication_year": row.get("publication_year", ""),
                "publication": row.get("publication", "")
        } for i, row in enumerate(page_data)] 

        return jsonify({
                "draw": draw,
                "recordsTotal": len(data),
                "recordsFiltered": len(filtered_data_sort),
                "data": results
            })
    
    with open("output.json", 'r', encoding='utf-8') as f:
            data = json.load(f)

    queries = set(item["query"].lower() for item in data if "query" in item)
    queries = list(queries)

    return render_template("results/results.html", queries=queries)

@app.route("/results/false-articles", methods=["GET", "POST"])
def false_articles():
    if request.method == "POST":
        selected_articles = request.form.getlist("selected_articles")
        #selected = ", ".join(selected_articles)

        with open("output.json", encoding='utf-8') as file:
            articles = json.load(file)
        
        # Filter data berdasarkan judul yang dipilih
        false_articles = [article for article in articles if article['title'] in selected_articles]

        #store into false database
        for article in false_articles:
            cur = mysql.connection.cursor()

            cur.execute("SELECT * FROM false_articles where title = %s", (article['title'],))
            result = cur.fetchone()

            if result:
                print("Item already in database: %s" % (article['title'],))
            else:
                cur.execute(
                    """INSERT INTO false_articles (query, author, title, title_url, cited_by_value, cited_by_url, publication_year, publication, thumbnail) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (article['query'],
                     article['author'],
                     article['title'],
                     article['title_url'],
                     article['cited_by_value'],
                     article['cited_by_url'],
                     article['publication_year'],
                     article['publication'],
                     article['thumbnail']
                     )
                )
                mysql.connection.commit()
                cur.close()

                # Hapus data yang sudah tersimpan dari output.json
                remaining_articles = [article for article in articles if article['title'] not in selected_articles]
        
                # Tulis ulang sisa artikel yang belum tersimpan ke dalam output.json
                with open("output.json", "w") as file:
                    json.dump(remaining_articles, file, indent=2)

        # Format string pengembalian untuk debug
        #return f"Received selected_articles: {false_articles}"
        return redirect(url_for('results'))
    
@app.route("/results/false-articles/view")
def false_articles_view():
    if request.args.get('draw'):
        draw = request.args.get('draw')
        start = int(request.args.get('start', 0))  # start point
        length = int(request.args.get('length', 10))  # number of records per page
        search_value = request.args.get('search[value]', '')  # search term
    
        # Ambil parameter sorting dari DataTables
        order_column = request.args.get('order[0][column]', '1')  # Default column is query
        order_dir = request.args.get('order[0][dir]', 'asc')  # Default direction is asc
    
        # Inisialisasi variabel untuk menyimpan arah sorting masing-masing kolom
        query_direction = 'asc'  # Default untuk query
        year_direction = 'asc'   # Default untuk publication_year
    
        # Logika sorting yang independen
        if order_column == "1":  # Jika user mengklik kolom query
            query_direction = order_dir
        
        elif order_column == "5":  # Jika user mengklik kolom publication_year
            year_direction = order_dir
            # Ambil sorting terakhir untuk query dari session jika ada
            prev_query_dir = request.args.get('prev_query_dir', 'asc')
            query_direction = prev_query_dir
    
        # Susun string ORDER BY
        order_by = f"query {query_direction}, publication_year {year_direction}"
        
        # Query with searching
        search_clause = f"WHERE query LIKE %s OR author LIKE %s OR title LIKE %s OR cited_by_value LIKE %s OR publication_year LIKE %s OR publication LIKE %s"
        search_param = f"%{search_value}%"
        
        # Main query with search, sorting, and pagination
        query = (
                f"SELECT * FROM false_articles {search_clause} ORDER BY {order_by} LIMIT %s OFFSET %s"
        )
        
        cur = mysql.connection.cursor()
        cur.execute(query, (search_param, search_param, search_param, search_param, search_param, search_param, length, start))
        false_articles = cur.fetchall()

        # Replace None values with '0'
        false_articles = [[0 if value is None else value for value in row] for row in false_articles]

        # Count total data
        cur.execute("SELECT COUNT(*) FROM false_articles")
        total_data = cur.fetchone()[0]

        # Count filtered data (based on search query)
        cur.execute(f"SELECT COUNT(*) FROM false_articles {search_clause}", 
                        (search_param, search_param, search_param, search_param, search_param, search_param))
        filtered_data = cur.fetchone()[0]

        cur.close()

        # Prepare the data in DataTables format
        results = [{
                "no": start + i + 1,
                "id_false_articles": row[0],
                "query": row[1],
                "author": row[2],
                "title": row[3],
                "title_url": row[4],
                "cited_by_value": row[5],
                "cited_by_url": row[6],
                "publication_year": row[7],
                "publication": row[8],
        } for i, row in enumerate(false_articles)]

        # Return the JSON response
        return jsonify({
                "draw": draw,
                "recordsTotal": total_data,
                "recordsFiltered": filtered_data,
                "data": results
        })

    return render_template("results/false_articles.html")

@app.route("/results/false-articles/delete/<int:id_false_articles>", methods=["GET", "POST"])
def false_articles_delete(id_false_articles):
    cur = mysql.connection.cursor()
    cur.execute("""SELECT query, author, title, title_url, cited_by_value, cited_by_url, publication_year, publication, thumbnail 
                FROM false_articles WHERE id_false_articles=  %s""", (id_false_articles,))
    false_articles = cur.fetchone()

    # Baca output.json saat ini
    with open('output.json', encoding='utf-8') as file:
        output_data = json.load(file)
        
    # Tambahkan data false_articles ke output.json
    publication = {
        'query': false_articles[0],
        'author': false_articles[1],
        'title': false_articles[2],
        'title_url': false_articles[3],
        'cited_by_value': false_articles[4],
        'cited_by_url': false_articles[5],
        'publication_year': false_articles[6],
        'publication': false_articles[7],
        'thumbnail': false_articles[8]
    }
    output_data.append(publication)

    # Tulis ulang data ke output.json
    with open('output.json', 'w', encoding='utf-8') as file:
        json.dump(output_data, file, indent=2)

    cur.execute("DELETE FROM false_articles WHERE id_false_articles = %s", (id_false_articles,))

    mysql.connection.commit()
    cur.close()

    flash('Publikasi berhasil dihapus dari list.', 'success')
    return redirect(url_for("false_articles_view"))

@app.route("/results/filtered-articles", methods=["GET", "POST"])
def filtered_articles():
    with open("output.json", encoding='utf-8') as file:
        articles = json.load(file)
    
    filtered_articles = [article for article in articles]
    for article in filtered_articles:
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM publikasi where title = %s", (article['title'],))
        result = cur.fetchone()

        if result:
            print("Item already in database: %s" % (article['title'],))
        else:
            publication_year = article['publication_year']
            if not publication_year or not publication_year.isdigit():
                publication_year = None
            cur.execute(
                """INSERT INTO publikasi (query, author, title, title_url, cited_by_value, cited_by_url, publication_year, publication, thumbnail) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (article['query'],
                article['author'],
                article['title'],
                article['title_url'],
                article['cited_by_value'],
                article['cited_by_url'],
                publication_year,
                article['publication'],
                article['thumbnail'],
                )
            )
            mysql.connection.commit()
            cur.close()
    return redirect(url_for('all_results'))
    


#@app.route("/all_results", methods=["GET", "POST"])
#def all_results():
#    page = request.args.get("page", default=1, type=int)
#    if request.method == "POST":
#        per_page = int(request.form.get("per_page", "10"))
#        return redirect(url_for('all_results', per_page=per_page, page=1))
#    else:
#        per_page = int(request.args.get("per_page", "10"))
#    offset = (page - 1) * per_page
#    sort_by_query = request.args.get("sortQuery", default="query", type=str)
#    sort_order_query = request.args.get("orderQuery", default="asc", type=str)
#    sort_by_year = request.args.get("sortYear", default="publication_year", type=str)
#    sort_order_year = request.args.get("orderYear", default="asc", type=str)
#    # Buat klausa ORDER BY untuk pengurutan
#    order_clause = f"ORDER BY {sort_by_query} {sort_order_query}, {sort_by_year} {sort_order_year}"
#    # Query data dari database dengan filter dan paginasi
#    query = (
#        f"SELECT * FROM publikasi {order_clause} LIMIT {per_page} OFFSET {offset}"
#    )
#    cur = mysql.connection.cursor()
#    cur.execute(query)
#    # cur.execute('''SELECT author, title, publication_year, link FROM crawlings''')
#    data = cur.fetchall()
#    # mengganti nilai None dengan '0'
#    data = [[0 if value is None else value for value in row] for row in data]
#    
#    # Hitung total data untuk paginasi
#    total_data_query = f"SELECT COUNT(*) FROM publikasi"
#    cur.execute(total_data_query)
#    total_data = cur.fetchone()[0]
#    # hitung total halaman
#    total_pages = (total_data // per_page) + (1 if total_data % per_page > 0 else 0)
#    cur.close()
#    return render_template(
#        "results/all_results.html",
#        values=data,
#        total_pages=total_pages,
#        current_page=page,
#        page=page,
#        per_page=per_page,
#        sort_by_query=sort_by_query,
#        sort_order_query=sort_order_query,
#        sort_by_year=sort_by_year,
#        sort_order_year=sort_order_year
#    )

@app.route("/all_results", methods=["GET", "POST"])
def all_results():
    try:
        if request.args.get('draw'):
            draw = request.args.get('draw')
            start = int(request.args.get('start', 0))  # start point
            length = int(request.args.get('length', 10))  # number of records per page
            search_value = request.args.get('search[value]', '')  # search term
            
            # Ambil parameter sorting dari DataTables
            order_column = request.args.get('order[0][column]', '1')  # Default column is query
            order_dir = request.args.get('order[0][dir]', 'asc')  # Default direction is asc
            
            # Inisialisasi variabel untuk menyimpan arah sorting masing-masing kolom
            query_direction = 'asc'  # Default untuk query
            year_direction = 'asc'   # Default untuk publication_year
            
            # Logika sorting yang independen
            if order_column == "1":  # Jika user mengklik kolom query
                query_direction = order_dir
                # year_direction tetap asc (default) atau mengikuti sorting terakhir
                
            elif order_column == "5":  # Jika user mengklik kolom publication_year
                year_direction = order_dir
                # Ambil sorting terakhir untuk query dari session jika ada
                prev_query_dir = request.args.get('prev_query_dir', 'asc')
                query_direction = prev_query_dir
            
            # Susun string ORDER BY
            order_by = f"query {query_direction}, publication_year {year_direction}"
            
            # Debug: Print sorting parameters
            print(f"Clicked Column: {order_column}")
            print(f"Click Direction: {order_dir}")
            print(f"Query Direction: {query_direction}")
            print(f"Year Direction: {year_direction}")
            print(f"Final ORDER BY: {order_by}")
            
            # Query with searching
            search_clause = f"WHERE query LIKE %s OR author LIKE %s OR title LIKE %s OR cited_by_value LIKE %s OR publication_year LIKE %s"
            search_param = f"%{search_value}%"

            # Main query with search, sorting, and pagination
            query = (
                f"SELECT * FROM publikasi {search_clause} ORDER BY {order_by} LIMIT %s OFFSET %s"
            )

            cur = mysql.connection.cursor()
            #print(f"Executing query: {query}")  # Tambahkan logging query
            cur.execute(query, (search_param, search_param, search_param, search_param, search_param, length, start))
            data = cur.fetchall()

            # Replace None values with '0'
            data = [[0 if value is None else value for value in row] for row in data]

            # Count total data
            cur.execute("SELECT COUNT(*) FROM publikasi")
            total_data = cur.fetchone()[0]

            # Count filtered data (based on search query)
            cur.execute(f"SELECT COUNT(*) FROM publikasi {search_clause}", 
                        (search_param, search_param, search_param, search_param, search_param))
            filtered_data = cur.fetchone()[0]

            cur.close()

            # Prepare the data in DataTables format
            results = [{
                "no": start + i + 1,
                "query": row[1],
                "author": row[2],
                "title": row[3],
                "title_url": row[4],
                "cited_by_value": row[5],
                "cited_by_url": row[6],
                "publication_year": row[7]
            } for i, row in enumerate(data)]

            # Return the JSON response
            return jsonify({
                "draw": draw,
                "recordsTotal": total_data,
                "recordsFiltered": filtered_data,
                "data": results
            })
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500
    
    # Default GET request for the template rendering
    return render_template("results/all_results.html")


@app.route("/results/preview_pdf")
def preview_pdf():
    with open("output.json", "r") as file_json:
        data_json = json.load(file_json)
        # Mengganti nilai None dengan 0
        for entry in data_json:
            for key, value in entry.items():
                if value is None:
                    entry[key] = 0

    # for entry in data_json:
    #   entry['title'] = entry['title'][0] if entry['title'] else ""

    rendered = render_template("template/template_pdf.html", data_json=data_json)

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    options = {
        "enable-local-file-access": None,
        "page-size": "A4",
        "page-width": "210mm",
        "page-height": "330mm",
        "margin-top": "20mm",
        "margin-right": "20mm",
        "margin-bottom": "20mm",
        "margin-left": "20mm",
    }
    pdf = pdfkit.from_string(rendered, False, configuration=config, options=options)

    return send_file(
        io.BytesIO(pdf),
        mimetype="application/pdf",
        as_attachment=False,
        download_name="preview.pdf",
    )

@app.route("/results/download/excel")
def download_excel():
    page = request.args.get("page", 1, type=int)

    with open("output.json", "r") as f:
        data = json.load(f)

    data_sorted = sorted(data, key=lambda x: x["query"])

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_sorted)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Set alignment and font for the header row
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Hasil_crawling.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/results/search-data-json", methods=["GET", "POST"])
def search_data_json():
    search_query = request.form.get("search_query")

    with open("output.json", "r") as f:
        data_json = json.load(f)

    # for entry in data_json:
    #   entry['title'] = entry['title'][0] if entry['title'] else ""

    search_results = [
        item
        for item in data_json
        if search_query.lower() in item["query"]
        or search_query.lower() in item["author"]
        or search_query.lower() in item["title"]
        or search_query.lower() in item["cited_by_value"]
        or search_query.lower() in item["publication_year"]
    ]

    return jsonify(search_results)

@app.route("/all_results/download/excel")
def all_results_download_excel():
    page = request.args.get("page", 1, type=int)

    cur = mysql.connection.cursor()
    cur.execute(
        """SELECT query, author, title, title_url, cited_by_value, cited_by_url, publication_year FROM publikasi"""
    )
    values = cur.fetchall()
    cur.close()

    # Convert the list of tuples to a list of dictionaries
    columns = [
        "query",
        "author",
        "title",
        "title_url",
        "cited_by_value",
        "cited_by_url",
        "publication_year",
    ]
    data_sorted = sorted(
        [dict(zip(columns, row)) for row in values], key=lambda x: x["query"].lower()
    )

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_sorted)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Set alignment and font for the header row
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Hasil_crawling.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/all_results/preview_pdf")
def all_results_preview_pdf():
    cur = mysql.connection.cursor()
    cur.execute(
        """SELECT query, author, title, title_url, cited_by_value, cited_by_url, publication_year FROM publikasi ORDER BY author ASC"""
    )
    data = cur.fetchall()

    # mengganti nilai None dengan '0'
    data = [[0 if value is None else value for value in row] for row in data]

    cur.close()

    # for entry in data_json:
    #   entry['title'] = entry['title'][0] if entry['title'] else ""

    rendered = render_template("template/template_pdf_database.html", database=data)

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    options = {
        "enable-local-file-access": None,
        "page-size": "A4",
        "page-width": "210mm",
        "page-height": "330mm",
        "margin-top": "20mm",
        "margin-right": "20mm",
        "margin-bottom": "20mm",
        "margin-left": "20mm",
    }
    pdf = pdfkit.from_string(rendered, False, configuration=config, options=options)

    return send_file(
        io.BytesIO(pdf),
        mimetype="application/pdf",
        as_attachment=False,
        download_name="preview.pdf",
    )


#ALL_RESULTS/PENCARIAN
#@app.route("/all_results/search-author", methods=["GET", "POST"])
#def all_results_search_author():
#    page = request.args.get("page", default=1, type=int)
#    author = request.args.get("query", "")
#    session['get_author']=author
#
#    if request.method == "POST":
#        per_page = int(request.form.get("per_page", "10"))
#        author = request.form.get("query", "")
#        return redirect(url_for('all_results_search_author', per_page=per_page, page=1, query=author))
#    else:
#        per_page = int(request.args.get("per_page", "10"))
#        author = request.args.get("query", "")
#    offset = (page - 1) * per_page
#
#    sort_by_query = request.args.get("sortQuery", default="query", type=str)
#    sort_order_query = request.args.get("orderQuery", default="asc", type=str)
#
#    sort_by_year = request.args.get("sortYear", default="publication_year", type=str)
#    sort_order_year = request.args.get("orderYear", default="asc", type=str)
#
#    # Buat klausa ORDER BY untuk pengurutan
#    order_clause = f"ORDER BY {sort_by_query} {sort_order_query}, {sort_by_year} {sort_order_year}"
#    
#    query = f"SELECT * FROM publikasi WHERE query = %s {order_clause} LIMIT %s OFFSET %s"
#
#    cur = mysql.connection.cursor()
#    cur.execute(query, (author, per_page, offset))
#    data = cur.fetchall()
#
#    # mengganti nilai None dengan '0'
#    data = [[0 if value is None else value for value in row] for row in data]
#
#    total_data_query = "SELECT COUNT(*) FROM publikasi WHERE query = %s"
#    cur.execute(total_data_query, (author,))
#    total_data = cur.fetchone()[0]
#
#    total_pages = (total_data // per_page) + (1 if total_data % per_page > 0 else 0)
#
#    return render_template(
#        "results/search_author_results.html",
#        search_value=data,
#        author=author,
#        page=page,
#        total_pages=total_pages,
#        per_page=per_page,
#        sort_by_query=sort_by_query,
#        sort_order_query=sort_order_query,
#        sort_by_year=sort_by_year,
#        sort_order_year=sort_order_year
#    )
@app.route("/all_results/search-author", methods=["GET", "POST"])
def all_results_search_author():
    try:    
        if request.method == "POST":
            author = request.form.get("query", "")
            session['author'] = author
            return redirect(url_for('all_results_search_author'))
        
        if request.args.get('draw'):
            author = session.get("author", "")
            print("Author:", author)
            
            #parameter datatables
            draw = request.args.get('draw')
            start = int(request.args.get('start', 0))
            length = int(request.args.get('length', 10))
            search_value = request.args.get('search[value]', '')

            order_column = request.args.get('order[0][column]', '5')  # Default column is year
            order_dir = request.args.get('order[0][dir]', 'asc')  # Default direction is asc

            # Inisialisasi variabel untuk menyimpan arah sorting masing-masing kolom
            year_direction = 'asc'   # Default untuk publication_year

            # Logika sorting yang independen
            if order_column == "5":  # Jika user mengklik kolom year
                year_direction = order_dir
            
            # Susun string ORDER BY
            order_by = f"publication_year {year_direction}"
    
            # Query with searching
            search_clause = "WHERE query = %s AND (author LIKE %s OR title LIKE %s OR cited_by_value LIKE %s OR publication_year LIKE %s)"
            search_param = f"%{search_value}%"
    
            query = f"SELECT * FROM publikasi {search_clause} ORDER BY {order_by} LIMIT %s OFFSET %s"
    
            cur = mysql.connection.cursor()
            cur.execute(query, (author, search_param, search_param, search_param, search_param, length, start))
            data = cur.fetchall()
    
            print("SQL Query:", query)
            print("Parameters:", (author, search_param, search_param, search_param, search_param, length, start))
    
            # mengganti nilai None dengan '0'
            data = [[0 if value is None else value for value in row] for row in data]
    
            # Count total data
            cur.execute("SELECT COUNT(*) FROM publikasi")
            total_data = cur.fetchone()[0]
    
            # Query untuk menghitung data yang difilter
            cur.execute(f"SELECT COUNT(*) FROM publikasi {search_clause}", 
                            (author, search_param, search_param, search_param, search_param))
            total_filtered_data = cur.fetchone()[0]
    
            results = [{
                "no": start + i + 1,
                "query": row[1],
                "author": row[2],
                "title": row[3],
                "title_url": row[4],
                "cited_by_value": row[5],
                "cited_by_url": row[6],
                "publication_year": row[7]
            } for i, row in enumerate(data)]
    
            cur.close()
            return jsonify({
                    "draw": draw,
                    "recordsTotal": total_data,
                    "recordsFiltered": total_filtered_data,
                    "data": results,
            })
        
        author = session.get("author", "")
        return render_template("results/search_author_results.html", author=author)
               
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/search-author", methods=["GET"])
def search_author():
    author = request.args.get("query", "").lower()

    cur = mysql.connection.cursor()
    if author:
        cur.execute(
            "SELECT query FROM publikasi WHERE query LIKE %s ORDER BY query ASC", ("%" + author + "%",)
        )
    else:
        cur.execute("SELECT query FROM publikasi ORDER BY query ASC")
    author = cur.fetchall()
    cur.close()

    results = list(set([result[0] for result in author]))
    results.sort()

    return jsonify(results)

@app.route("/all_results/search-author/download/excel", methods=["GET", "POST"])
def search_author_download_excel():
    author = request.args.get("query", "")

    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT query, author, title, title_url, cited_by_value, cited_by_url, publication_year FROM publikasi WHERE query = %s",
        (author,),
    )
    values = cur.fetchall()
    cur.close()

    columns = [
        "query",
        "author",
        "title",
        "title_url",
        "cited_by_value",
        "cited_by_url",
        "publication_year",
    ]
    data_sorted = sorted(
        [dict(zip(columns, row)) for row in values], key=lambda x: x["author"].lower()
    )

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_sorted)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Set alignment and font for the header row
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Hasil_crawling.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/all_results/search-author/preview_pdf", methods=["GET"])
def search_author_preview_pdf():
    author = request.args.get("query", "")

    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT query, author, title, title_url, cited_by_value, cited_by_url, publication_year FROM publikasi WHERE query = %s ORDER BY author ASC",
        (author,),
    )
    data = cur.fetchall()

    # mengganti nilai None dengan '0'
    data = [[0 if value is None else value for value in row] for row in data]

    cur.close()

    # for entry in data_json:
    #   entry['title'] = entry['title'][0] if entry['title'] else ""

    rendered = render_template("template/template_pdf_search.html", query=data)

    config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

    options = {
        "enable-local-file-access": None,
        "page-size": "A4",
        "page-width": "210mm",
        "page-height": "330mm",
        "margin-top": "20mm",
        "margin-right": "20mm",
        "margin-bottom": "20mm",
        "margin-left": "20mm",
    }
    pdf = pdfkit.from_string(rendered, False, configuration=config, options=options)

    return send_file(
        io.BytesIO(pdf),
        mimetype="application/pdf",
        as_attachment=False,
        download_name="preview.pdf",
    )

@app.route("/all_results/search_bar", methods=["GET", "POST"])
def search_bar():
    page = request.args.get("page", default=1, type=int)
    search_query = request.args.get("search", "").lower()

    if request.method == "POST":
        per_page = int(request.form.get("per_page", "10"))
        search_query = request.args.get("search", "").lower()
        return redirect(url_for('search_bar', per_page=per_page, page=1, search_query=search_query))
    else:
        per_page = int(request.args.get("per_page", "10"))
        search_query = request.args.get("search", "").lower()
    offset = (page - 1) * per_page

    sort_by_query = request.args.get("sortQuery", default="query", type=str)
    sort_order_query = request.args.get("orderQuery", default="asc", type=str)

    sort_by_cite = request.args.get("sortCite", default="cited_by_value", type=str)
    sort_order_cite = request.args.get("orderCite", default="asc", type=str)

    sort_by_year = request.args.get("sortYear", default="publication_year", type=str)
    sort_order_year = request.args.get("orderYear", default="asc", type=str)

    per_page = 10
    offset = (page - 1) * per_page

    # klausa WHERE untuk filter author
    search_clause = f"LOWER(query) LIKE %s OR LOWER(author) LIKE %s OR LOWER(title) LIKE %s OR publication_year LIKE %s"
    search_values = (
        "%" + search_query + "%",
        "%" + search_query + "%",
        "%" + search_query + "%",
        "%" + search_query + "%",
    )

    # Buat klausa ORDER BY untuk pengurutan
    order_clause = f"ORDER BY {sort_by_query} {sort_order_query}, {sort_by_cite} {sort_order_cite}, {sort_by_year} {sort_order_year}"

    # Gabungkan klausa ORDER BY dengan klausa WHERE
    query = f"SELECT * FROM publikasi WHERE {search_clause} {order_clause} LIMIT {per_page} OFFSET {offset}"

    cur = mysql.connection.cursor()
    cur.execute(query, search_values)
    data = cur.fetchall()

    # mengganti nilai None dengan '0'
    data = [[0 if value is None else value for value in row] for row in data]


    # Hitung total data untuk paginasi
    total_data_query = f"SELECT COUNT(*) FROM publikasi WHERE {search_clause}"
    cur.execute(total_data_query, search_values)
    total_data = cur.fetchone()[0]

    total_pages = (total_data // per_page) + (1 if total_data % per_page > 0 else 0)

    return render_template(
        "results/search_bar_results.html",
        search_value=data,
        search_query=search_query,
        total_pages=total_pages,
        page=page,
        per_page=per_page,
        sort_by_query=sort_by_query,
        sort_order_query=sort_order_query,
        sort_by_cite=sort_by_cite,
        sort_order_cite=sort_order_cite,
        sort_by_year=sort_by_year,
        sort_order_year=sort_order_year,
    )


#DOSEN
#@app.route("/dosen")
#def dosen():
#    page = request.args.get("page", 1, type=int)
#    cur = mysql.connection.cursor()
#    cur.execute("""SELECT * FROM dosen""")
#    data_dosen = cur.fetchall()
#    cur.close()
#    # Tentukan jumlah data per halaman
#    per_page = 10
#    total = len(data_dosen)
#    total_pages = math.ceil(total / per_page)
#    # Tentukan batasan data untuk halaman saat ini
#    start = (page - 1) * per_page
#    end = start + per_page
#    paginated_data = data_dosen[start:end]
#    return render_template(
#        "/dosen/dosen.html",
#        dosen=paginated_data,
#        page=page,
#        per_page=per_page,
#        total_pages=total_pages,
#    )

@app.route("/dosen")
def dosen():
    if request.args.get('draw'):
        draw = request.args.get('draw')
        start = int(request.args.get('start', 0))  # start point
        length = int(request.args.get('length', 10))  # number of records per page
        search_value = request.args.get('search[value]', '')  # search term
        
        # Ambil parameter sorting dari DataTables
        order_column = request.args.get('order[0][column]', '1')  # Default column is nama_dosen
        order_dir = request.args.get('order[0][dir]', 'asc')  # Default direction is asc

        name_direction = 'asc'
        prodi_direction = 'asc'
        
        if order_column == "1":  # Jika user mengklik kolom nama dosen
            name_direction = order_dir

        elif order_column == "2":  # Jika user mengklik kolom publication_year
            prodi_direction = order_dir
            # Ambil sorting terakhir untuk nama dosen dari session jika ada
            prev_query_dir = request.args.get('prev_query_dir', 'asc')
            name_direction = prev_query_dir
        
        #order_by = f"nama_dosen {name_direction}, prodi {prodi_direction}"

        order_by = (
            f"nama_dosen {name_direction}, "
            f"SUBSTRING_INDEX(prodi, ' ', 1) {prodi_direction}, "  # Sort by level (e.g., S-1, S-2)
            f"SUBSTRING_INDEX(prodi, ' ', -1) {prodi_direction}"  # Sort by program name within each level
        )

        # Query with searching
        search_clause = f"WHERE nama_dosen LIKE %s OR prodi LIKE %s"
        search_param = f"%{search_value}%"

        # Main query with search, sorting, and pagination
        query = (
                f"SELECT * FROM dosen {search_clause} ORDER BY {order_by} LIMIT %s OFFSET %s"
            )

        try:
            cur = mysql.connection.cursor()
            cur.execute(query, (search_param, search_param, length, start))
            data_dosen = cur.fetchall()

            # Count total data
            cur.execute("SELECT COUNT(*) FROM dosen")
            total_data = cur.fetchone()[0]

            # Count filtered data (based on search query)
            cur.execute(f"SELECT COUNT(*) FROM dosen {search_clause}", 
                        (search_param, search_param))
            filtered_data = cur.fetchone()[0]

            cur.close()

            # Prepare the data in DataTables format
            results = [{
                "no": start + i + 1,
                "id_dosen": row[0],
                "nama_dosen": row[1],
                "prodi": row[2]
            } for i, row in enumerate(data_dosen)]

            # Return the JSON response
            return jsonify({
                "draw": draw,
                "recordsTotal": total_data,
                "recordsFiltered": filtered_data,
                "data": results
            })
        
        except Exception as e:
            # Log error for debugging
            print("Error fetching data:", str(e))
            return jsonify({"error": "Data fetch error"}), 500
    
    return render_template("/dosen/dosen.html")


@app.route("/dosen/add", methods=["GET", "POST"])
def dosen_add():
    return render_template("dosen/add.html")

@app.route("/dosen/add/submit", methods=["GET", "POST"])
def dosen_add_submit():
    nama_dosen = request.form["nama_dosen"]
    prodi = request.form["prodi"]

    cur = mysql.connection.cursor()
    cur.execute(
        "INSERT INTO dosen (nama_dosen, prodi) VALUES (%s, %s)", (nama_dosen, prodi)
    )
    mysql.connection.commit()
    cur.close()

    flash('Dosen berhasil ditambahkan.', 'success')
    return redirect(url_for("dosen"))

@app.route("/dosen/delete/<int:id_dosen>", methods=["POST", "DELETE"])
def dosen_delete(id_dosen):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM dosen WHERE id_dosen = %s", (id_dosen,))
    mysql.connection.commit()
    cur.close()

    flash('Dosen berhasil dihapus.', 'success')
    return redirect(url_for("dosen"))

@app.route("/dosen/update/form/<int:id_dosen>", methods=["GET", "POST"])
def dosen_update_form(id_dosen):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM dosen WHERE id_dosen = %s", (id_dosen,))
    dosen = cur.fetchone()
    cur.close()

    return render_template("dosen/update-form.html", dosen=dosen)

@app.route("/dosen/update/<int:id_dosen>", methods=["POST", "GET"])
def dosen_update(id_dosen):
    nama_dosen = request.form["nama_dosen"]
    prodi = request.form["prodi"]

    cur = mysql.connection.cursor()
    cur.execute(
        "UPDATE dosen SET nama_dosen = %s, prodi = %s WHERE id_dosen = %s",
        (nama_dosen, prodi, id_dosen),
    )
    mysql.connection.commit()
    cur.close()

    flash('Dosen berhasil diperbarui.', 'success')
    return redirect(url_for("dosen"))

@app.route("/dosen/import", methods=["GET", "POST"])
def dosen_import():
    if request.method == "POST":
        if "file" not in request.files:
            return "No file part"

        file = request.files["file"]

        if file.filename == "":
            return "No selected file"

        df = pd.read_excel(file)

        cur = mysql.connection.cursor()
        for index, row in df.iterrows():
            cur.execute(
                "INSERT INTO dosen (nama_dosen, prodi) VALUES (%s, %s)",
                (row["nama_dosen"], row["prodi"]),
            )
        mysql.connection.commit()
        cur.close()

        flash('Dosen berhasil diimport.', 'success')
        return redirect(url_for("dosen"))
    return redirect(url_for("dosen"))



#AUTOCOMPLETE
@app.route("/search-query", methods=["GET"])
def search_query():
    query = request.args.get("nama", "")

    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT nama_dosen as query FROM dosen WHERE nama_dosen LIKE %s",
        ("%" + query + "%",),
    )
    results = cur.fetchall()
    cur.close()
    return jsonify([result[0] for result in results])


#OTHERS
@app.route("/search_bar_author", methods=["GET", "POST"])
def search_bar_author():
    page = request.args.get("page", default=1, type=int)
    search_query = request.args.get("search", "").lower()
    query_param = request.args.get("author", "").lower()
    author = session.get("get_author", "-")

    sort_by_query = request.args.get("sortQuery", default="query", type=str)
    sort_order_query = request.args.get("orderQuery", default="asc", type=str)

    sort_by_cite = request.args.get("sortCite", default="cited_by_value", type=str)
    sort_order_cite = request.args.get("orderCite", default="asc", type=str)

    sort_by_year = request.args.get("sortYear", default="publication_year", type=str)
    sort_order_year = request.args.get("orderYear", default="asc", type=str)

    per_page = 10
    offset = (page - 1) * per_page

    # klausa WHERE untuk filter author
    search_clause = f"(LOWER(author) LIKE %s OR LOWER(title) LIKE %s OR publication_year LIKE %s) AND LOWER(query) LIKE %s"
    search_values = (
        "%" + search_query + "%",
        "%" + search_query + "%",
        "%" + search_query + "%",
        "%" + query_param + "%"
    )

    # Buat klausa ORDER BY untuk pengurutan
    order_clause = f"ORDER BY {sort_by_query} {sort_order_query}, {sort_by_cite} {sort_order_cite}, {sort_by_year} {sort_order_year}"

    # Gabungkan klausa ORDER BY dengan klausa WHERE
    query = f"SELECT * FROM publikasi WHERE {search_clause} {order_clause} LIMIT {per_page} OFFSET {offset}"

    cur = mysql.connection.cursor()
    cur.execute(query, search_values)
    data = cur.fetchall()

    # mengganti nilai None dengan '-'
    sanitized_values = []
    for row in data:
        sanitized_row = [(value if value is not None else '-') for value in row]
        sanitized_values.append(sanitized_row)

    # Hitung total data untuk paginasi
    total_data_query = f"SELECT COUNT(*) FROM publikasi WHERE {search_clause}"
    cur.execute(total_data_query, search_values)
    total_data = cur.fetchone()[0]

    total_pages = (total_data // per_page) + (1 if total_data % per_page > 0 else 0)

    return render_template(
        "results/search_bar_author.html",
        author=author,
        search_value=sanitized_values,
        search_query=search_query,
        total_pages=total_pages,
        page=page,
        per_page=per_page,
        sort_by_query=sort_by_query,
        sort_order_query=sort_order_query,
        sort_by_cite=sort_by_cite,
        sort_order_cite=sort_order_cite,
        sort_by_year=sort_by_year,
        sort_order_year=sort_order_year,
    )

@app.route("/author-crawling", methods=["POST", "GET"])
def author_crawling():
    spider_name = "dosenCrawler"

    subprocess.run(["scrapy", "crawl", spider_name, "-o", "output_dosen.json"])

    with open("output.json") as items_file:
        return render_template("author.html", author_output=items_file.read())

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=FLASK_PORT)
